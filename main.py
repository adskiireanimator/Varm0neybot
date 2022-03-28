import telebot
import openpyxl
import datetime
from telebot import types

with open('config.txt', 'r') as s:
    token = s.readline()
bot = telebot.TeleBot(token)
name = ''
age = 0
id_user = ''
category = ''
shop = ''
value = 0


@bot.message_handler(content_types=['text'])
def menu(message):
    # Основная фунция, которая обрабатывает разные сообщения и распределяет дальнейшие пути развития
    # (20 значущих строчек кода, границу проходит)
    wb = openpyxl.load_workbook('varmoney2.xlsx')  # открываем книгу
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    if not (str(message.from_user.id) in wb.sheetnames):  # смотрим есть ли у нас такой юзер в базе данных
        if message.text == "/reg":
            markup.add(types.KeyboardButton('Варечка'))
            bot.send_message(message.from_user.id, 'Как тебя зовут?', reply_markup=None)
            bot.register_next_step_handler(message, get_name)
        else:
            markup.add(types.KeyboardButton('/reg'))
            bot.send_message(message.from_user.id, 'Привет, меня зовут VarMoneyBot,' +
                             ' я помогу тебе лучше анализировать расходы \n\n\nнапиши мне /reg', reply_markup=markup)
    else:
        ws = wb[str(message.from_user.id)]  # открываем лист этого юзера
        btn1 = types.KeyboardButton('Отчет за день')
        btn2 = types.KeyboardButton('Отчет за месяц')
        btn3 = types.KeyboardButton('Добавить покупочку')
        btn4 = types.KeyboardButton('Смотреть по категориям')
        btn5 = types.KeyboardButton('Смотреть по магазинам')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.from_user.id, f"{ws['A1'].value}, что ты хочешь? \n", reply_markup=markup)
        bot.register_next_step_handler(message, choose_action)  # запускаем следующие шаги


def choose_action(message):
    # Функия меню, позволяет пользователю выбрать действие в меню
    # (30 значущих строчек кода, границу проходят)
    er_id = message.from_user.id  # берём id юзера
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)  # создаём клавиатуру
    if message.text == 'Отчет за день':  # смотрим что именно за сообщение
        markup.add(types.KeyboardButton('/menu'))  # добавляем кнопку
        need_date = str(datetime.datetime.now()).split(' ')[0]  # форматируем дату
        report = day_report(er_id, need_date)  # запускаеми функцию, которая посчитает результат
        bot.send_message(er_id, report + f'\n что-нибудь еще? /menu', reply_markup=markup)  # пишем пользователю
    elif message.text == 'Отчет за месяц':  # аналогично
        markup.add(types.KeyboardButton('/menu'))
        need_date = str(datetime.datetime.now()).split(' ')[0].split('-')[1]
        report = month_report(er_id, need_date)
        bot.send_message(er_id, report + f'\n что-нибудь еще? /menu', reply_markup=markup)
    elif message.text == 'Добавить покупочку':  # тоже самое, только вместо результата запускаем следующий шаг
        btn1 = types.KeyboardButton('По категориям')  # здесь мы создали кнопки отдельно
        btn2 = types.KeyboardButton('По магазинам')
        btn3 = types.KeyboardButton('Анонимно')
        markup.add(btn1, btn2, btn3)
        bot.send_message(er_id, f'Как вы хотите зарегистрировать покупку? \n', reply_markup=markup)
        bot.register_next_step_handler(message, purchase_value)
    elif message.text == 'Смотреть по категориям':  # аналогично
        btn1 = types.KeyboardButton('За день')
        btn2 = types.KeyboardButton('За месяц')
        markup.add(btn1, btn2)
        bot.send_message(er_id, 'выберите промежуток', reply_markup=markup)
        bot.register_next_step_handler(message, category_checker)
    elif message.text == 'Смотреть по магазинам':  # аналогично
        btn1 = types.KeyboardButton('за день')
        btn2 = types.KeyboardButton('за месяц')
        markup.add(btn1, btn2)
        bot.send_message(er_id, 'Выберите промежуток', reply_markup=markup)
        bot.register_next_step_handler(message, shop_checker)
    else:  # проверка на то что пользователь напишет что-то своё
        bot.send_message(er_id, "Я не знаю таких команд")


def category_checker(message):
    # функция, которая проверяет за какой промежуток времени хотят посмотреть и спршивает уже именно категорию
    # запускает следующий шаг
    # (16 значущих строчек, границу проходят)
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    if message.text == 'За день':  # узнаем промежуток времени
        categories = criteria_searcher('category', str(message.from_user.id))
        # записываем все категории, которые есть у пользователя
        answer = 'Какую категорию вы хотите выбрать? '
        for x in categories:
            markup.add(types.KeyboardButton(x))  # добавляем в клавиатуру каждую категорию человека
        bot.send_message(message.from_user.id, answer, reply_markup=markup)  # предагаем пользователю выбрать категорию
        bot.register_next_step_handler(message, category_report_day)
    elif message.text == 'За месяц':  # аналогично
        categories = criteria_searcher('category', str(message.from_user.id))
        answer = 'Какую категорию вы хотите выбрать? '
        for x in categories:
            markup.add(types.KeyboardButton(x))
        bot.send_message(message.from_user.id, answer, reply_markup=markup)
        bot.register_next_step_handler(message, category_report_month)
    else:  # проверка если напишет не то что нужно
        bot.send_message(message.from_user.id, "такой катгеории нет", reply_markup=None)


def category_report_month(message):
    # принимает в себя категорию, парсит таблицу и выдаёт результат
    # (15 значущих строчек)
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(types.KeyboardButton('Меню'))
    categories = criteria_searcher('category', str(message.from_user.id))
    date = str(datetime.datetime.now()).split(' ')[0].split('-')[1]
    if message.text in categories:  # проверка нахождения такой категории
        wb = openpyxl.load_workbook('varmoney2.xlsx')
        ws = wb[str(message.from_user.id)]
        report = 0
        for i in range(2, ws.max_column + 1):
            if str(ws.cell(column=i, row=1).value).split(' ')[0].split('-')[1] == date and \
                    ws.cell(column=i, row=3).value == message.text:  # проверка по датек и самой категории
                report += int(ws.cell(column=i, row=2).value)  # если всё хорошо то мы просто добавляем в наш отчёт
        report = str(report)
        bot.send_message(message.from_user.id, report + '\n что-нибудь еще?', reply_markup=markup)  # пишем юзеру
    else:  # проверка если он что-то своё написал
        bot.send_message(message.from_user.id, "Ничего не нашёл", reply_markup=markup)


def category_report_day(message):
    # принимает в себя категорию, парсит таблицу и выдаёт результат
    # (15 значущих строчек)
    # также как для месяцв
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(types.KeyboardButton('Меню'))
    categories = criteria_searcher('category', str(message.from_user.id))
    date = str(datetime.datetime.now()).split(' ')[0]
    if message.text in categories:
        wb = openpyxl.load_workbook('varmoney2.xlsx')
        ws = wb[str(message.from_user.id)]
        report = 0
        for i in range(2, ws.max_column + 1):
            if str(ws.cell(column=i, row=1).value).split(' ')[0] == date and \
                    ws.cell(column=i, row=3).value == message.text:
                report += int(ws.cell(column=i, row=2).value)
        report = str(report)
        bot.send_message(message.from_user.id, report + '\n что-нибудь еще?', reply_markup=markup)
    else:
        bot.send_message(message.from_user.id, "Ничего не нашёл", reply_markup=markup)


def shop_report_month(message):
    # принимает в себя магазин, парсит таблицу и выдаёт результат
    # (15 значущих строчек)
    # также как для категорий, прям точь в точь
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(types.KeyboardButton('Меню'))
    categories = criteria_searcher('shop', str(message.from_user.id))
    date = str(datetime.datetime.now()).split(' ')[0].split('-')[1]
    if message.text in categories:
        wb = openpyxl.load_workbook('varmoney2.xlsx')
        ws = wb[str(message.from_user.id)]
        report = 0
        for i in range(2, ws.max_column + 1):
            if str(ws.cell(column=i, row=1).value).split(' ')[0].split('-')[1] == date and \
                    ws.cell(column=i, row=3).value == message.text:
                report += int(ws.cell(column=i, row=2).value)
        report = str(report)
        bot.send_message(message.from_user.id, report + '\n что-нибудь еще?', reply_markup=markup)
    else:
        bot.send_message(message.from_user.id, "Ничего не нашёл", reply_markup=markup)


def shop_report_day(message):
    # принимает в себя магазин, парсит таблицу и выдаёт результат
    # (15 значущих строчек)
    # также как для категорий, прям точь в точь
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(types.KeyboardButton('Меню'))
    categories = criteria_searcher('shop', str(message.from_user.id))
    date = str(datetime.datetime.now()).split(' ')[0]
    if message.text in categories:
        wb = openpyxl.load_workbook('varmoney2.xlsx')
        ws = wb[str(message.from_user.id)]
        report = 0
        for i in range(2, ws.max_column + 1):
            if str(ws.cell(column=i, row=1).value).split(' ')[0] == date and \
                    ws.cell(column=i, row=3).value == message.text:
                report += int(ws.cell(column=i, row=2).value)
        report = str(report)
        bot.send_message(message.from_user.id, report + '\n что-нибудь еще?', reply_markup=markup)
    else:
        bot.send_message(message.from_user.id, "Ничего не нашёл", reply_markup=markup)


def shop_checker(message):
    # функция, которая проверяет за какой промежуток времени хотят посмотреть и спршивает уже именно магазин
    # запускает следующий шаг
    # (16 значущих строчек, границу проходят)
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    if message.text == 'За день':
        categories = criteria_searcher('shop', str(message.from_user.id))
        answer = 'Какой магазин вы хотите выбрать? '
        for x in categories:
            markup.add(types.KeyboardButton(x))
        bot.send_message(message.from_user.id, answer, reply_markup=markup)
        bot.register_next_step_handler(message, shop_report_day)
    elif message.text == 'За месяц':
        categories = criteria_searcher('shop', str(message.from_user.id))
        answer = 'Какой магазин вы хотите выбрать? '
        for x in categories:
            markup.add(types.KeyboardButton(x))
        bot.send_message(message.from_user.id, answer, reply_markup=markup)
        bot.register_next_step_handler(message, shop_report_month)
    else:
        bot.send_message(message.from_user.id, "такого магазина нет", reply_markup=None)


def day_report(day_user_id, date):
    # Функция, которая принимает id юзера и возвращает его расходы за 1 день
    # (6 значущих строчек кода, границу проходят)
    wb = openpyxl.load_workbook('varmoney2.xlsx')
    ws = wb[str(day_user_id)]
    report = 0
    for i in range(2, ws.max_column + 1):
        if str(ws.cell(column=i, row=1).value).split(' ')[0] == date:
            report += int(ws.cell(column=i, row=2).value)  # просто парсим табличку и если дата совпадает с датой
    return str(report)


def month_report(id_month_user, date):
    # Функция, которая принимает id юзера и возвращает его расходы за 1 месяц
    # (6 значущих строчек кода, границу проходят)
    wb = openpyxl.load_workbook('varmoney2.xlsx')
    ws = wb[str(id_month_user)]
    report = 0
    for i in range(2, ws.max_column + 1):
        if str(ws.cell(column=i, row=1).value).split(' ')[0].split('-')[1] == date:
            report += int(ws.cell(column=i, row=2).value)  # также что и с днем только немного по другому проверяем дату
    return str(report)


def purchase_value(message):
    # Функия, которая распределяет способы добавления расходов юзера по остальным функциям
    # (19 значущих строчек кода, границу проходят)
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    if message.text == 'Анонимно':  # смотрим какой именно вид добавления расходов он выбрал
        markup.add(types.KeyboardButton('0'))
        bot.send_message(message.from_user.id,
                         "Введите нужную сумму...", reply_markup=None)  # предлагаем ему выбрать сумму расходов
        bot.register_next_step_handler(message, anonim_purchase)  # запускаем следующий шаг
    elif message.text == 'По категориям':
        categories = criteria_searcher('category', str(message.from_user.id))
        answer = 'Какую категорию вы хотите выбрать?\n(Вы можете написать свой вариант)'
        for x in categories:
            markup.add(types.KeyboardButton(x))
        bot.send_message(message.from_user.id, answer, reply_markup=markup)
        # предлагаем выбрать одну из категорий, либо написать самому
        bot.register_next_step_handler(message, category_add)
    elif message.text == 'По магазинам':
        shops = criteria_searcher('shop', str(message.from_user.id))
        answer = 'Какой магазин вы хотите выбрать?\n(Вы можете написать свой вариант)'
        for x in shops:
            markup.add(types.KeyboardButton(x))
        bot.send_message(message.from_user.id, answer, reply_markup=markup)
        # предлагаем выбрать один из магазинов, либо написать самому
        bot.register_next_step_handler(message, shop_add)


def criteria_searcher(criterion, user_id):
    # функция по данному ей критерию находит внутри пользователя все его возможные либо магазины, либо категории
    # (6 значущих строчек кода, границу проходят)
    criteria = set()
    wb = openpyxl.load_workbook('varmoney2.xlsx')
    ws = wb[user_id]
    for i in range(2, ws.max_column + 1):
        if ws.cell(column=i, row=4).value == criterion:
            criteria.add(ws.cell(column=i, row=3).value)
    return list(criteria)


def category_add(message):
    # фунция получает категорию в сообщении, записывает в глобальную переменную и запускает следующий шаг
    # (4 значущие строчки кода, границу проходят)
    global category
    category = message.text  # просто записываем в глоабльную переменную имя категории
    bot.send_message(message.from_user.id, "Введите нужную сумму...", reply_markup=None)
    bot.register_next_step_handler(message, category_purchase)


def category_purchase(message):
    # функция получает сумму, котрую хочет зарегистрировать человек,
    # записывает в файл сумму и категорию из глобальной переменой и возвращает в меню
    # (15 значущих строчек кода, границу проходят)
    global category
    category_id = str(message.from_user.id)
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    try:
        markup.add(types.KeyboardButton('меню'))
        value_category = int(message.text)
        wb = openpyxl.load_workbook('varmoney2.xlsx')
        ws = wb[category_id]
        data = str(datetime.datetime.now()).split(' ')[0]
        ws.cell(column=ws.max_column + 1, row=1).value = data  # записываем дату когда он зарегистрировал покупуку
        ws.cell(column=ws.max_column, row=2).value = value_category  # регистрируем стоимость
        ws.cell(column=ws.max_column, row=3).value = category  # ставим нужную категорию
        ws.cell(column=ws.max_column, row=4).value = 'category'  # и записываем флажок категории
        category = ''  # обнуляем глобальную переменную
        wb.save('varmoney2.xlsx')
        bot.send_message(message.from_user.id, 'Данные сохранены) \n', reply_markup=markup)
    except ValueError:  # проверка на то что она написал не число, а например какое-то слово
        bot.send_message(message.from_user.id, 'Пиши цифрами', reply_markup=None)


def shop_add(message):
    # фунция получает магазин в сообщении, записывает в глобальную переменную и запускает следующий шаг
    # (4 значущие строчки кода, границу проходят)
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    global shop
    shop = message.text  # просто записываем в глоабльную переменную имя магазина
    markup.add(types.KeyboardButton('0'))
    bot.send_message(message.from_user.id, "Введите нужную сумму...", reply_markup=None)
    bot.register_next_step_handler(message, shop_purchase)


def shop_purchase(message):
    # функция получает сумму, котрую хочет зарегистрировать человек,
    # записывает в файл сумму и магазин из глобальной переменной и возвращает в меню
    # (15 значущих строчек кода, границу проходят)
    global shop
    shop_id = str(message.from_user.id)
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    try:
        shop_value = int(message.text)
        markup.add(types.KeyboardButton('Меню'))
        wb = openpyxl.load_workbook('varmoney2.xlsx')
        ws = wb[shop_id]
        data = str(datetime.datetime.now()).split(' ')[0]
        ws.cell(column=ws.max_column + 1, row=1).value = data  # записываем дату когда он зарегистрировал покупуку
        ws.cell(column=ws.max_column, row=2).value = shop_value  # регистрируем стоимость
        ws.cell(column=ws.max_column, row=3).value = shop  # ставим нужный магазин
        ws.cell(column=ws.max_column, row=4).value = 'shop'  # и записываем флажок магазина
        shop = ''
        wb.save('varmoney2.xlsx')
        bot.send_message(message.from_user.id, 'Данные сохранены) \n', reply_markup=markup)
    except ValueError:  # опять же проверяем на то что это число
        bot.send_message(message.from_user.id, 'Пиши цифрами', reply_markup=None)


def anonim_purchase(message):
    # получает сумму и записывает в файл без категорий и магазинов
    # (11 значущих строчек кода, границу проходят)
    anonim_id = str(message.from_user.id)
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(types.KeyboardButton('Меню'))
    try:
        anonim_value = int(message.text)
        need_date = str(datetime.datetime.now()).split(' ')[0]
        wb = openpyxl.load_workbook('varmoney2.xlsx')
        ws = wb[anonim_id]
        ws.cell(column=ws.max_column + 1, row=1).value = need_date  # записываем дату
        ws.cell(column=ws.max_column, row=2).value = anonim_value  # записываем стоимость
        wb.save('varmoney2.xlsx')  # записываем в книгу без флажка так как это анонимные расходы
        bot.send_message(message.from_user.id, 'Данные сохранены) \n', reply_markup=markup)
    except ValueError:  # проверка на число
        bot.send_message(message.from_user.id, 'Пиши цифрами', reply_markup=markup)


def get_name(message):
    # фунция получает ник человека в сообщении, записывает в глобальную переменную и запускает следующий шаг
    # (4 значущие строчки кода, границу проходят)
    global name
    name = message.text  # запись в глобальную перемену
    bot.send_message(message.from_user.id, "Сколько тебе лет?")
    bot.register_next_step_handler(message, get_age_and_id)


def get_age_and_id(message):
    # записывает в новую учетную запись ник, возраст человека и id его телеграмма
    # (18 значущих строчек кода, границу проходят)
    global age
    global name
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(types.KeyboardButton('Меню'))
    while age == 0:  # не даём уйти с регистрации до тех пор он ведёт своё возраст
        try:
            age = int(message.text)
        except ValueError:
            bot.send_message(message.from_user.id, 'Пиши цифрами')
    id_for_age = message.from_user.id
    wb = openpyxl.load_workbook('./varmoney2.xlsx')
    wb.create_sheet(str(id_for_age))  # создаём новый лист в екселе с названием его id в телеграмме
    sheet = wb[str(id_for_age)]
    sheet['A1'] = name
    sheet['A2'] = age
    wb.save('varmoney2.xlsx')
    name = ''  # обнуляем глоабльные переменные
    age = 0
    bot.send_message(id_for_age, 'Спaсибо за регистрацию \n', reply_markup=markup)


bot.polling(none_stop=True, interval=0)
